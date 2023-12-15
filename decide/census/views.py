import csv
import json
import openpyxl
import xml.etree.ElementTree as ET
from django.contrib.auth.decorators import user_passes_test

from django.db.utils import IntegrityError
from django.shortcuts import render
from django.utils.decorators import method_decorator
from django.views import View
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_protect
from rest_framework import generics
from rest_framework.response import Response
from django.http import JsonResponse
from django.http import HttpResponseRedirect
from openpyxl import load_workbook
from rest_framework.status import (
    HTTP_201_CREATED as ST_201,
    HTTP_204_NO_CONTENT as ST_204,
    HTTP_409_CONFLICT as ST_409
)

from datetime import datetime
from django.utils import timezone
from django.shortcuts import get_object_or_404
from base.perms import UserIsStaff
from .models import Census
from store.models import Vote


class CensusCreate(generics.ListCreateAPIView):
    permission_classes = (UserIsStaff,)

    def create(self, request, *args, **kwargs):
        voting_id = request.data.get('voting_id')
        voters = request.data.get('voters')

        try:
            census = Census(voting_id=voting_id, voter_id=voters[0])
            census.save()

            if (timezone.now() - census.creation_date).days >= 7:
                census.status = 'Desactivado'
            else:
                census.status = 'Activo'

            census.has_voted = Vote.objects.filter(voting_id=voting_id, voter_id=voters[0]).exists()
            census.save()

        except IntegrityError:
            return Response('Error trying to create census', status=ST_409)

        return Response('Census created', status=ST_201)

    def list(self, request, *args, **kwargs):
        voting_id = request.GET.get('voting_id')
        creation_date = request.GET.get('creation_date')

        queryset = Census.objects.filter(voting_id=voting_id)

        if creation_date:
            queryset = queryset.filter(creation_date__gte=creation_date)

        data = []
        for census in queryset:
            status = 'Desactivado' if (timezone.now() - census.creation_date).days >= 7 else 'Activo'
            has_voted = Vote.objects.filter(voting_id=voting_id, voter_id=census.voter_id).exists()
            data.append({
                'voter_id': census.voter_id,
                'creation_date': census.creation_date.strftime('%d-%m-%Y %H:%M:%S'),
                'status': status,
                'total_voters': Census.objects.filter(voting_id=voting_id).count(),
                'has_voted': has_voted,
            })

        return Response(data)


class CensusDetail(generics.RetrieveDestroyAPIView):

    def destroy(self, request, voting_id, *args, **kwargs):
        voters = request.data.get('voters')
        census = Census.objects.filter(voting_id=voting_id, voter_id__in=voters)
        census.delete()
        return Response('Voters deleted from census', status=ST_204)

    def retrieve(self, request, voting_id, *args, **kwargs):
        voter = request.GET.get('voter_id')
        census = get_object_or_404(Census, voting_id=voting_id, voter_id=voter)
        return Response({
            'voting_id': census.voting_id,
            'voter_id': census.voter_id,
            'creation_date': census.creation_date.strftime('%d-%m-%Y %H:%M:%S'),
            'status': 'Desactivado' if (timezone.now() - census.creation_date).days >= 7 else 'Activo',
            'total_voters': Census.objects.filter(voting_id=voting_id).count(),
        })


# ---------

@method_decorator(user_passes_test(lambda u: u.is_authenticated and u.is_staff), name='dispatch')
@method_decorator(csrf_protect, name='dispatch')
class CensusExportView(View):
    template_name = 'export_census.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request):
        export_format = request.POST.get('export_format')

        url_mapping = {
        'csv': 'export_census_csv/',
        'json': 'export_census_json/',
        'xlsx': 'export_census_xlsx/',
        'xml': 'export_census_xml/',
    }

        if export_format in url_mapping:
            return HttpResponseRedirect(url_mapping[export_format])
        else:
            return render(request, 'export_census.html', {'error_message': 'Export format not valid.'})


@method_decorator(user_passes_test(lambda u: u.is_authenticated and u.is_staff), name='dispatch')
@method_decorator(csrf_protect, name='dispatch')
class CensusImportView(View):
    template_name = 'import_census.html'
    SUPPORTED_CONTENT_TYPES = ['text/csv', 'application/json',
                               'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']

    def process_file(self, file, content_type):
        if content_type == 'text/csv':
            decoded_file = file.read().decode('utf-8').splitlines()
            reader = csv.reader(decoded_file)
            next(reader, None)
        elif content_type == 'application/json':
            data = json.loads(file.read().decode('utf-8'))
            reader = (item.values() for item in data)
        elif content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            workbook = load_workbook(file, read_only=True)
            sheet = workbook.active
            reader = sheet.iter_rows(min_row=2, values_only=True)
        else:
            return None, JsonResponse({'error': 'Unsupported file format'}, status=400)

        return reader, None

    def create_census_object(self, row):
        voting_id, voter_id, creation_date_str, additional_info = row

        if additional_info == None:
            additional_info = ''

        if voting_id is None or voter_id is None or creation_date_str is None:
            raise ValueError('Incomplete data in row')

        if Census.objects.filter(voting_id=voting_id, voter_id=voter_id).exists():
            raise ValueError('Duplicate data found')

        return Census(
            voting_id=voting_id,
            voter_id=voter_id,
            creation_date=timezone.now(),
            additional_info=additional_info
        )

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        if request.method == 'POST':
            file = request.FILES.get('file')
            if file:
                content_type = file.content_type
                reader, response = self.process_file(file, content_type)
                if response:
                    return response

                try:
                    for row in reader:
                        if isinstance(row, list):
                            voting_id = row[0]
                            voter_id = row[1]
                        else:
                            row_list = list(row)
                            voting_id = row_list[0]
                            voter_id = row_list[1]

                        if (voting_id != None and voter_id != None):
                            census_object = self.create_census_object(row)
                            census_object.save()

                    return JsonResponse({'message': 'Census imported successfully'}, status=201)

                except ValueError as ve:
                    return JsonResponse({'error': str(ve)}, status=400)

                except Exception as e:
                    return JsonResponse({'error': f'Error trying to create census: {str(e)}'}, status=500)

            return JsonResponse({'error': 'Invalid or no file provided'}, status=400)

        return JsonResponse({'error': 'Invalid request method'}, status=405)


@method_decorator(user_passes_test(lambda u: u.is_authenticated and u.is_staff), name='dispatch')
@method_decorator(csrf_protect, name='dispatch')
class ExportCensusToCSV(View):

    def get(self, request):
        census_data = Census.objects.all()

        response = self.export_to_csv(census_data)

        return response

    def export_to_csv(self, census_data):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="census.csv"'

        writer = csv.writer(response)
        writer.writerow(['Voting ID', 'Voter ID', 'Creation Date', 'Additional Info'])

        for census in census_data:
            writer.writerow([census.voting_id, census.voter_id, census.creation_date, census.additional_info])

        return response


@method_decorator(user_passes_test(lambda u: u.is_authenticated and u.is_staff), name='dispatch')
@method_decorator(csrf_protect, name='dispatch')
class ExportCensusToJSON(View):

    def get(self, request):
        census_data = Census.objects.all()
        response = self.export_to_json(census_data)
        return response

    def export_to_json(self, census_data):
        export_data = []
        for census in census_data:
            export_data.append({
                'voting_id': census.voting_id,
                'voter_id': census.voter_id,
                'creation_date': census.creation_date.strftime('%Y-%m-%d %H:%M:%S'),
                'additional_info': census.additional_info,
            })

        json_data = json.dumps(export_data, indent=2, default=str)

        response = HttpResponse(json_data, content_type='application/json')
        response['Content-Disposition'] = 'attachment; filename="census.json"'

        return response


@method_decorator(user_passes_test(lambda u: u.is_authenticated and u.is_staff), name='dispatch')
@method_decorator(csrf_protect, name='dispatch')
class ExportCensusToXLSX(View):

    def get(self, request):
        census_data = Census.objects.all()
        response = self.export_to_excel(census_data)
        return response

    def export_to_excel(self, census_data):
        if not census_data:
            return HttpResponse('No hay datos para exportar a Excel.', status=204)

        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        headers = [field.name for field in Census._meta.get_fields() if field.name != 'id']
        worksheet.append(headers)

        for census in census_data:
            formatted_date = census.creation_date.strftime('%Y-%m-%d %H:%M:%S')
            data_row = [getattr(census, field) if field != 'creation_date' else formatted_date for field in headers]
            worksheet.append(data_row)

        file_name = f"census_export_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'

        workbook.save(response)
        workbook.close()

        return response

@method_decorator(user_passes_test(lambda u: u.is_authenticated and u.is_staff), name='dispatch')
@method_decorator(csrf_protect, name='dispatch')
class ExportCensusToXML(View):
    def get(self, request):
        census_data = Census.objects.all()
        response = self.export_to_xml(census_data)
        return response

    def export_to_xml(self, census_data):
        root = ET.Element("CensusData")

        for census in census_data:
            census_element = ET.SubElement(root, "Census")
            ET.SubElement(census_element, "VotingID").text = str(census.voting_id)
            ET.SubElement(census_element, "VoterID").text = str(census.voter_id)
            ET.SubElement(census_element, "CreationDate").text = census.creation_date.strftime('%Y-%m-%d %H:%M:%S')
            ET.SubElement(census_element, "AdditionalInfo").text = census.additional_info

        xml_data = ET.tostring(root, encoding="utf-8")

        response = HttpResponse(xml_data, content_type="application/xml")
        response["Content-Disposition"] = 'attachment; filename="census.xml"'

        return response