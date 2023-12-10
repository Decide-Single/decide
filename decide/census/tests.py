import json
import os
import tempfile
from datetime import datetime
from datetime import timedelta

import openpyxl
from base.tests import BaseTestCase

from django.contrib.staticfiles.testing import StaticLiveServerTestCase
from django.core.files.uploadedfile import SimpleUploadedFile
from django.contrib.auth.models import User
from django.http import HttpResponseRedirect
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl.workbook import Workbook
from operator import attrgetter
from io import BytesIO
from selenium import webdriver
from selenium.webdriver.common.by import By

from store.models import Vote

from .models import Census



class CensusFilterTestCase(BaseTestCase):
    
    def setUp(self):
        super().setUp()

        current_date = timezone.now()

        time_deltas = [timezone.timedelta(weeks=i) for i in range(4)]

        self.censuses = [
            Census(voting_id=1, voter_id=1, additional_info="Prueba 1", creation_date=current_date - time_deltas[0]),
            Census(voting_id=1, voter_id=2, creation_date=current_date - time_deltas[1]),
            Census(voting_id=1, voter_id=3, creation_date=current_date - time_deltas[2]),
            Census(voting_id=1, voter_id=4, creation_date=current_date - time_deltas[3]),
        ]

        for census in self.censuses:
            census.save()

        self.votes = [
            Vote(voter_id=1, voting_id=1, a="Valor_a_1", b="Valor_b_1"),
        ]

        for vote in self.votes:
            vote.save()
            
    def tearDown(self):
        super().tearDown()
        self.census = None

    def test_list_census(self):
        self.login()
        self.assertEqual(Census.objects.count(), 4)

    def test_filter_census_desactivate(self):
        self.login()
        all_census_objects = Census.objects.all()
        filtered_census = [census for census in all_census_objects if census.get_status() == 'Desactivado']
        self.assertEqual(len(filtered_census), 3)

    def test_filter_census_activate(self):
        self.login()
        all_census_objects = Census.objects.all()
        filtered_census = [census for census in all_census_objects if census.get_status() == 'Activo']
        self.assertEqual(len(filtered_census), 1)

    def test_filter_census_total_voters(self):
        self.login()
        all_census_objects = Census.objects.all()
        filtered_census = [census for census in all_census_objects if census.get_total_voters() == 4]
        self.assertEqual(len(filtered_census), 4)

    def test_filter_census_has_not_voted(self):
        self.login()
        all_census_objects = Census.objects.all()
        filtered_census = [census for census in all_census_objects if not census.has_voted()]
        self.assertEqual(len(filtered_census), 3)

    def test_filter_census_has_voted(self):
        self.login()
        all_census_objects = Census.objects.all()
        filtered_census = [census for census in all_census_objects if census.has_voted()]
        self.assertEqual(len(filtered_census), 1)

class CensusTestCase(BaseTestCase):

    def setUp(self):
        super().setUp()
        self.census = Census(voting_id=1, voter_id=1)
        self.census.save()

    def tearDown(self):
        super().tearDown()
        self.census = None

    def test_check_vote_permissions(self):
        
        response = self.client.get('/census/{}/?voter_id={}'.format(1, 2), format='json')
        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.json().get('detail'), 'Not found.')

        response = self.client.get('/census/{}/?voter_id={}'.format(1, 1), format='json')
        self.assertEqual(response.status_code, 200)

    def test_list_voting(self):
        response = self.client.get('/census/?voting_id={}'.format(1), format='json')
        self.assertEqual(response.status_code, 401)

        self.login(user='noadmin')
        response = self.client.get('/census/?voting_id={}'.format(1), format='json')
        self.assertEqual(response.status_code, 403)

        self.login()
        response = self.client.get('/census/?voting_id={}'.format(1), format='json')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()[0].get('voter_id'), 1)

    def test_add_new_voters_conflict(self):
        data = {'voting_id': 1, 'voters': [1]}
        response = self.client.post('/census/', data, format='json')
        self.assertEqual(response.status_code, 401)

        self.login(user='noadmin')
        response = self.client.post('/census/', data, format='json')
        self.assertEqual(response.status_code, 403)

        self.login()
        response = self.client.post('/census/', data, format='json')
        self.assertEqual(response.status_code, 409)

    def test_add_new_voters(self):
        data = {'voting_id': 2, 'voters': [1,2]}
        response = self.client.post('/census/', data, format='json')
        self.assertEqual(response.status_code, 401)

        self.login(user='noadmin')
        response = self.client.post('/census/', data, format='json')
        self.assertEqual(response.status_code, 403)

        self.login()
        response = self.client.post('/census/', data, format='json')
        self.assertEqual(response.status_code, 201)

        self.assertEqual(len(data.get('voters')), Census.objects.count())

    def test_destroy_voter(self):
        data = {'voters': [1]}
        response = self.client.delete('/census/{}/'.format(1), data, format='json')
        self.assertEqual(response.status_code, 204)
        self.assertEqual(0, Census.objects.count())

class CensusTest(StaticLiveServerTestCase):
    def setUp(self):
        #Load base test functionality for decide
        self.base = BaseTestCase()
        self.base.setUp()

        options = webdriver.ChromeOptions()
        options.headless = True
        self.driver = webdriver.Chrome(options=options)

        super().setUp()

    def tearDown(self):
        super().tearDown()
        self.driver.quit()

        self.base.tearDown()
    
    def createCensusSuccess(self):
        self.cleaner.get(self.live_server_url+"/admin/login/?next=/admin/")
        self.cleaner.set_window_size(1280, 720)

        self.cleaner.find_element(By.ID, "id_username").click()
        self.cleaner.find_element(By.ID, "id_username").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").click()
        self.cleaner.find_element(By.ID, "id_password").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").send_keys("Keys.ENTER")

        self.cleaner.get(self.live_server_url+"/admin/census/census/add")
        now = datetime.now()
        self.cleaner.find_element(By.ID, "id_voting_id").click()
        self.cleaner.find_element(By.ID, "id_voting_id").send_keys(now.strftime("%m%d%M%S"))
        self.cleaner.find_element(By.ID, "id_voter_id").click()
        self.cleaner.find_element(By.ID, "id_voter_id").send_keys(now.strftime("%m%d%M%S"))
        self.cleaner.find_element(By.NAME, "_save").click()

        self.assertTrue(self.cleaner.current_url == self.live_server_url+"/admin/census/census")

    def createCensusEmptyError(self):
        self.cleaner.get(self.live_server_url+"/admin/login/?next=/admin/")
        self.cleaner.set_window_size(1280, 720)

        self.cleaner.find_element(By.ID, "id_username").click()
        self.cleaner.find_element(By.ID, "id_username").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").click()
        self.cleaner.find_element(By.ID, "id_password").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").send_keys("Keys.ENTER")

        self.cleaner.get(self.live_server_url+"/admin/census/census/add")

        self.cleaner.find_element(By.NAME, "_save").click()

        self.assertTrue(self.cleaner.find_element_by_xpath('/html/body/div/div[3]/div/div[1]/div/form/div/p').text == 'Please correct the errors below.')
        self.assertTrue(self.cleaner.current_url == self.live_server_url+"/admin/census/census/add")

    def createCensusValueError(self):
        self.cleaner.get(self.live_server_url+"/admin/login/?next=/admin/")
        self.cleaner.set_window_size(1280, 720)

        self.cleaner.find_element(By.ID, "id_username").click()
        self.cleaner.find_element(By.ID, "id_username").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").click()
        self.cleaner.find_element(By.ID, "id_password").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").send_keys("Keys.ENTER")

        self.cleaner.get(self.live_server_url+"/admin/census/census/add")
        now = datetime.now()
        self.cleaner.find_element(By.ID, "id_voting_id").click()
        self.cleaner.find_element(By.ID, "id_voting_id").send_keys('64654654654654')
        self.cleaner.find_element(By.ID, "id_voter_id").click()
        self.cleaner.find_element(By.ID, "id_voter_id").send_keys('64654654654654')
        self.cleaner.find_element(By.NAME, "_save").click()

        self.assertTrue(self.cleaner.find_element_by_xpath('/html/body/div/div[3]/div/div[1]/div/form/div/p').text ==
                        'Please correct the errors below.')
        self.assertTrue(self.cleaner.current_url == self.live_server_url+"/admin/census/census/add")


class CensusImportViewTest(TestCase):

    def setUp(self):
        self.admin_user = User.objects.create_user(
            username='admin',
            password='admin_password',
            is_staff=True,
        )

    def tearDown(self):
        Census.objects.all().delete()

    def login_as_admin(self):
        self.client.force_login(self.admin_user)

    def logout(self):
        self.client.logout()

    def _assert_census_data(self, census_objects):
        census_objects = sorted(census_objects, key=attrgetter('voting_id'))

        self.assertEqual(census_objects[0].voting_id, 1)
        self.assertEqual(census_objects[0].voter_id, 1)
        self.assertEqual(census_objects[0].additional_info, 'Info1')

        self.assertEqual(census_objects[1].voting_id, 2)
        self.assertEqual(census_objects[1].voter_id, 2)
        self.assertEqual(census_objects[1].additional_info, 'Info2')

    def test_import_csv_success(self):
        self.login_as_admin()
        csv_file = SimpleUploadedFile("test.csv", b"voting_id,voter_id,creation_date,additional_info\n1,1,"
            b"22023-11-28 11:47:12.015914+00:00,Info1\n2,2,2023-11-28 11:47:12.015914+00:00,Info2", content_type="text/csv")
        url = reverse('import_census')
        response = self.client.post(url, {'file': csv_file}, format='multipart')

        self.assertEqual(response.status_code, 201)
        self.assertEqual(Census.objects.count(), 2)
        self._assert_census_data(Census.objects.all())
        self.logout()

    def test_import_json_success(self):
        self.login_as_admin()
        json_data = [{'voting_id': 1, 'voter_id': 1, 'creation_date': '2023-11-28 11:47:12.015914+00:00', 'additional_info': 'Info1'},
                     {'voting_id': 2, 'voter_id': 2, 'creation_date': '2023-11-28 11:47:12.015914+00:00', 'additional_info': 'Info2'}]
        json_file = SimpleUploadedFile("test.json", json.dumps(json_data).encode(), content_type="application/json")
        response = self.client.post(reverse('import_census'), {'file': json_file}, format='multipart')

        self.assertEqual(response.status_code, 201)
        self.assertEqual(Census.objects.count(), 2)
        self._assert_census_data(Census.objects.all())
        self.logout()

    def test_import_excel_success(self):
        self.login_as_admin()
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['voting_id', 'voter_id', 'creation_date', 'additional_info'])
        sheet.append([1, 1, '2023-11-28 11:47:12.015914+00:00', 'Info1'])
        sheet.append([2, 2, '2023-11-28 11:47:12.015914+00:00', 'Info2'])
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        excel_file = SimpleUploadedFile("test.xlsx", excel_buffer.read(),
                                        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response = self.client.post(reverse('import_census'), {'file': excel_file}, format='multipart')

        self.assertEqual(response.status_code, 201)
        self.assertEqual(Census.objects.count(), 2)
        self._assert_census_data(Census.objects.all())
        self.logout()

    def test_import_csv_failure(self):
        self.login_as_admin()
        csv_file = SimpleUploadedFile("test_failure.csv",
                                      b"voting_id,voter_id,creation_date,additional_info\n1,1,2023-11-28 11:47:12.015914+00:00",
                                      content_type="text/csv")
        url = reverse('import_census')
        response = self.client.post(url, {'file': csv_file}, format='multipart')

        self.assertEqual(response.status_code, 400)
        self.assertEqual(Census.objects.count(), 0)
        self.logout()

    def test_import_json_failure(self):
        self.login_as_admin()
        json_data = [{'voting_id': 1, 'voter_id': 1, 'creation_date': '2023-11-28 11:47:12.015914+00:00'}]
        json_file = SimpleUploadedFile("test_failure.json", json.dumps(json_data).encode(),
                                       content_type="application/json")
        response = self.client.post(reverse('import_census'), {'file': json_file}, format='multipart')

        self.assertEqual(response.status_code, 400)
        self.assertEqual(Census.objects.count(), 0)
        self.logout()

    def test_import_excel_failure(self):
        self.login_as_admin()
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['voting_id', 'voter_id', 'creation_date', 'additional_info'])
        sheet.append([1, 1, '2023-11-28 11:47:12.015914+00:00'])
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        excel_file = SimpleUploadedFile("test_failure.xlsx", excel_buffer.read(),
                                        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response = self.client.post(reverse('import_census'), {'file': excel_file}, format='multipart')

        self.assertEqual(response.status_code, 400)
        self.assertEqual(Census.objects.count(), 0)
        self.logout()

    def test_import_csv_duplicate_data(self):
        self.login_as_admin()
        csv_file = SimpleUploadedFile("test_duplicate.csv",
                                      b"voting_id,voter_id,creation_date,additional_info\n1,1,2023-11-28 11:47:12.015914+00:00,"
                                      b"Info1\n1,1,2023-11-28 11:47:12.015914+00:00,Info1",
                                      content_type="text/csv")
        url = reverse('import_census')
        response = self.client.post(url, {'file': csv_file}, format='multipart')

        self.assertEqual(response.status_code, 400)
        self.assertEqual(Census.objects.count(), 1)
        self.logout()

    def test_import_json_duplicate_data(self):
        self.login_as_admin()
        json_data = [{'voting_id': 1, 'voter_id': 1, 'creation_date': '2023-11-28 11:47:12.015914+00:00',
                      'additional_info': 'Info1'},
                     {'voting_id': 1, 'voter_id': 1, 'creation_date': '2023-11-28 11:47:12.015914+00:00',
                      'additional_info': 'Info1'}]
        json_file = SimpleUploadedFile("test_duplicate.json", json.dumps(json_data).encode(),
                                       content_type="application/json")
        response = self.client.post(reverse('import_census'), {'file': json_file}, format='multipart')

        self.assertEqual(response.status_code, 400)
        self.assertEqual(Census.objects.count(), 1)
        self.logout()

    def test_import_excel_duplicate_data(self):
        self.login_as_admin()
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['voting_id', 'voter_id', 'creation_date', 'additional_info'])
        sheet.append([1, 1, '2023-11-28 11:47:12.015914+00:00', 'Info1'])
        sheet.append([1, 1, '2023-11-28 11:47:12.015914+00:00', 'Info1'])
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        excel_file = SimpleUploadedFile("test_duplicate.xlsx", excel_buffer.read(),
                                        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response = self.client.post(reverse('import_census'), {'file': excel_file}, format='multipart')

        self.assertEqual(response.status_code, 400)
        self.assertEqual(Census.objects.count(), 1)
        self.logout()

    def test_invalid_file_format(self):
        self.login_as_admin()
        invalid_file = SimpleUploadedFile("test.txt", b"Invalid file content", content_type="text/plain")
        response = self.client.post(reverse('import_census'), {'file': invalid_file}, format='multipart')

        self.assertEqual(response.status_code, 400)
        self.assertEqual(Census.objects.count(), 0)
        self.assertEqual(response.json(), {'error': 'Unsupported file format'})
        self.logout()

    def test_non_admin_access(self):
        self.client.login(username='testuser', password='testpassword')
        url = reverse('import_census')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(Census.objects.count(), 0)
        self.client.logout()


class BaseExportTestCase(TestCase):
    def setUp(self):
        self.admin_user = User.objects.create_user(
            username='admin',
            password='admin_password',
            is_staff=True,
        )

        self.census_data = [
            {
                'voting_id': 1,
                'voter_id': 1,
                'creation_date': timezone.now(),
                'additional_info': 'Test Info 1',
            },
            {
                'voting_id': 2,
                'voter_id': 2,
                'creation_date': timezone.now() - timedelta(days=1),
                'additional_info': 'Test Info 2',
            },
            {
                'voting_id': 3,
                'voter_id': 3,
                'creation_date': timezone.now() - timedelta(days=2),
                'additional_info': 'Test Info 3',
            },
            {
                'voting_id': 4,
                'voter_id': 4,
                'creation_date': timezone.now() - timedelta(days=3),
                'additional_info': 'Test Info 4',
            },
            {
                'voting_id': 5,
                'voter_id': 5,
                'creation_date': timezone.now() - timedelta(days=4),
                'additional_info': 'Test Info 5',
            },
        ]

        self.census_instances = [Census.objects.create(**data) for data in self.census_data]

    def login_as_admin(self):
        self.client.force_login(self.admin_user)

    def logout(self):
        self.client.logout()


class ExportCensusToCSVTest(BaseExportTestCase):

    def test_export_headers_to_csv(self):
        self.login_as_admin()

        url = reverse('export_census_to_csv')
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)

        response_lines = response.content.decode('utf-8').splitlines()

        expected_headers = ['Voting ID', 'Voter ID', 'Creation Date', 'Additional Info']
        self.assertEqual(response_lines[0].split(','), expected_headers)

    def test_export_data_to_csv(self):
        self.login_as_admin()

        url = reverse('export_census_to_csv')

        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)

        response_lines = response.content.decode('utf-8').splitlines()

        with tempfile.NamedTemporaryFile(mode='w+', suffix='.csv', delete=False) as temp_file:
            response = self.client.get(url)
            temp_file.write(response.content.decode('utf-8'))

        try:
            self.assertEqual(response.status_code, 200)

            with open(temp_file.name, 'r') as file:
                file_content = file.read().splitlines()

            for i, census in enumerate(self.census_data):
                if i + 1 < len(response_lines):
                    self.assert_exported_data_matches_census(response_lines[i + 1], census)

                if i + 1 < len(file_content):
                    self.assert_exported_data_matches_census(file_content[i + 1], census)

        finally:
            os.remove(temp_file.name)

    def assert_exported_data_matches_census(self, actual_data, census):
        expected_data = [
            str(census['voting_id']),
            str(census['voter_id']),
            census['creation_date'].strftime('%Y-%m-%d %H:%M:%S'),
            census['additional_info']
        ]
        actual_data = actual_data.split(',')

        for j in [2]:
            self.assertEqual(actual_data[j].split('.')[0], expected_data[j].split('.')[0])

        self.assertEqual(actual_data[:2] + actual_data[3:], expected_data[:2] + expected_data[3:])


class ExportCensusToJSONTest(BaseExportTestCase):

    def test_export_list_to_json(self):
        self.login_as_admin()

        url = reverse('export_census_to_json')

        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)

        response_data = json.loads(response.content.decode('utf-8'))

        self.assertIsInstance(response_data, list)

        self.assertEqual(len(response_data), len(self.census_instances))

        for i, census_data in enumerate(response_data):
            if i < len(self.census_instances):
                self.assert_exported_census_data_matches_instance(census_data, self.census_instances[i])

    def test_exported_json_to_file(self):
        self.login_as_admin()

        url = reverse('export_census_to_json')

        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)

        response_data = json.loads(response.content.decode('utf-8'))

        with tempfile.NamedTemporaryFile(suffix='.json', delete=False, mode='w+b') as temp_file:
            temp_file.write(json.dumps(response_data, indent=2, default=str).encode('utf-8'))
            temp_file_path = temp_file.name

        self.assertTrue(os.path.exists(temp_file_path))

        with open(temp_file_path, 'r', encoding='utf-8') as temp_file:
            saved_data = json.load(temp_file)

        for i, saved_census_data in enumerate(saved_data):
            if i < len(self.census_instances):
                self.assert_exported_census_data_matches_instance(saved_census_data, self.census_instances[i])

        os.remove(temp_file_path)

    def assert_exported_census_data_matches_instance(self, exported_data, census_instance):
        self.assertEqual(exported_data['voting_id'], census_instance.voting_id)
        self.assertEqual(exported_data['voter_id'], census_instance.voter_id)
        self.assertEqual(exported_data['creation_date'], census_instance.creation_date.strftime('%Y-%m-%d %H:%M:%S'))
        self.assertEqual(exported_data['additional_info'], census_instance.additional_info)

        expected_keys = ['voting_id', 'voter_id', 'creation_date', 'additional_info']
        self.assertCountEqual(exported_data.keys(), expected_keys)

        self.assertRegex(exported_data['creation_date'], r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}')


class ExportCensusToXLSXTest(BaseExportTestCase):

    def test_export_headers_to_excel(self):
        self.login_as_admin()

        url = reverse('export_census_to_xlsx')

        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)

        self.assertIn('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', response['Content-Type'])

        self.assertIn('Content-Disposition', response)

        expected_filename = f"census_export_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        self.assertIn(f'filename="{expected_filename}"', response['Content-Disposition'])

        self.assertNotEqual(response.content, b'')

        if not Census.objects.exists():
            self.assertEqual(response.status_code, 204)

        try:
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                temp_file.write(response.content)

            workbook = openpyxl.load_workbook(temp_file.name)
            worksheet = workbook.active

            self.assertEqual(worksheet.title, 'Sheet')

            expected_headers = ['voting_id', 'voter_id', 'creation_date', 'additional_info']

            for col_num, header in enumerate(expected_headers, start=1):
                self.assertEqual(worksheet.cell(row=1, column=col_num).value, header)

        finally:
            temp_file.close()

    def test_export_data_to_excel(self):
        self.login_as_admin()

        url = reverse('export_census_to_xlsx')

        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)

        self.assertIn('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', response['Content-Type'])

        self.assertIn('Content-Disposition', response)

        expected_filename = f"census_export_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        self.assertIn(f'filename="{expected_filename}"', response['Content-Disposition'])

        self.assertNotEqual(response.content, b'')

        if not Census.objects.exists():
            self.assertEqual(response.status_code, 204)

        try:
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                temp_file.write(response.content)

            workbook = openpyxl.load_workbook(temp_file.name)
            worksheet = workbook.active

            self.assertEqual(worksheet.title, 'Sheet')

            self.assert_excel_headers(worksheet)

            self.assert_excel_data(worksheet)

        finally:
            temp_file.close()

    def assert_excel_headers(self, worksheet):
        expected_headers = ['voting_id', 'voter_id', 'creation_date', 'additional_info']

        for col_num, header in enumerate(expected_headers, start=1):
            self.assertEqual(worksheet.cell(row=1, column=col_num).value, header)

    def assert_excel_data(self, worksheet):
        for row_num, census_instance in enumerate(self.census_instances, start=2):
            self.assertEqual(worksheet.cell(row=row_num, column=1).value, census_instance.voting_id)
            self.assertEqual(worksheet.cell(row=row_num, column=2).value, census_instance.voter_id)
            self.assertEqual(worksheet.cell(row=row_num, column=3).value,
                             census_instance.creation_date.strftime('%Y-%m-%d %H:%M:%S'))
            self.assertEqual(worksheet.cell(row=row_num, column=4).value, census_instance.additional_info)



class CensusExportViewTest(BaseExportTestCase):

    def test_export_census_view_requires_admin_login(self):
        response = self.client.get(reverse('export_census'))
        self.assertEqual(response.status_code, 302)

    def test_export_census_to_csv_requires_admin_login(self):
        response = self.client.get(reverse('export_census_to_csv'))
        self.assertEqual(response.status_code, 302)

    def test_export_census_to_json_requires_admin_login(self):
        response = self.client.get(reverse('export_census_to_json'))
        self.assertEqual(response.status_code, 302)

    def test_export_census_to_xlsx_requires_admin_login(self):
        response = self.client.get(reverse('export_census_to_xlsx'))
        self.assertEqual(response.status_code, 302)

    def test_get_export_census_view_as_admin(self):
        self.login_as_admin()
        response = self.client.get(reverse('export_census'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'export_census.html')

    def test_post_valid_format_as_admin(self):
        self.login_as_admin()
        data = {'export_format': 'csv'}
        response = self.client.post(reverse('export_census'), data)
        expected_url = reverse('export_census_to_csv')
        self.assertIsInstance(response, HttpResponseRedirect)
        self.assertRedirects(response, expected_url)

    def test_post_invalid_format_as_admin(self):
        self.login_as_admin()
        data = {'export_format': 'invalid_format'}
        response = self.client.post(reverse('export_census'), data)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'export_census.html')
        self.assertIn('error_message', response.context)
        self.assertEqual(response.context['error_message'], 'Export format not valid.')

    def test_post_missing_format_as_admin(self):
        self.login_as_admin()
        response = self.client.post(reverse('export_census'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'export_census.html')
        self.assertIn('error_message', response.context)
        self.assertEqual(response.context['error_message'], 'Export format not valid.')

